"""Report output exporters.

XLSX via openpyxl; PDF via the shared xhtml2pdf renderer already used for label
PDFs (``core.tasks.labels._html_to_pdf_bytes``), which carries an SSRF-safe link
callback so user-authored report templates can't fetch remote/internal resources.
Both take the already-compiled grid (headers + rows keyed by translated header)
or rendered HTML, so they stay format-agnostic across report types.
"""
import io

XLSX_MIME = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
PDF_MIME = 'application/pdf'


def report_xlsx_bytes(headers, rows, sheet_title='Report'):
    """Render the report grid (headers + rows) into an .xlsx workbook (bytes)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    # Excel caps sheet names at 31 chars and forbids []:*?/\
    safe_title = ''.join(c for c in (sheet_title or 'Report') if c not in '[]:*?/\\')[:31] or 'Report'
    ws.title = safe_title

    ws.append(list(headers))
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for r in rows:
        ws.append([r.get(h, '-') for h in headers])

    for col_idx, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(40, len(str(h)) + 4))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def report_pdf_bytes(rendered_html):
    """Render already-compiled report HTML into PDF bytes via the shared
    xhtml2pdf renderer (same engine + SSRF-safe link callback as label PDFs)."""
    # inline import: reuse the label PDF renderer without a core.reports -> core.tasks
    # import at module load (and keep xhtml2pdf an on-demand dependency).
    from core.tasks.labels import _html_to_pdf_bytes
    return _html_to_pdf_bytes(rendered_html)
