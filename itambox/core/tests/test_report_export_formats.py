"""Tests for the report PDF + XLSX export formats (core/reports/exporters.py)."""
import io

from django.test import TestCase
from django.utils.translation import gettext as _

from core.tests.mixins import TenantTestMixin
from core.reports.exporters import report_xlsx_bytes, report_pdf_bytes


class ReportExporterUnitTests(TestCase):
    def test_xlsx_is_a_valid_workbook_with_headers_and_rows(self):
        headers = ['Asset Tag', 'Cost']
        rows = [{'Asset Tag': 'EXP-1', 'Cost': '€1,000.00'},
                {'Asset Tag': 'EXP-2', 'Cost': '-'}]
        data = report_xlsx_bytes(headers, rows, sheet_title='My Report')
        # XLSX is a zip container.
        self.assertEqual(data[:2], b'PK')
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        self.assertEqual([c.value for c in ws[1]], headers)
        self.assertEqual(ws.cell(row=2, column=1).value, 'EXP-1')
        self.assertEqual(ws.cell(row=2, column=2).value, '€1,000.00')
        self.assertEqual(ws.max_row, 3)

    def test_pdf_export_returns_pdf_bytes(self):
        html = "<html><body><h1>Hello Report</h1><table><tr><td>Row</td></tr></table></body></html>"
        data = report_pdf_bytes(html)
        self.assertTrue(data.startswith(b'%PDF'))
        self.assertGreater(len(data), 500)


class ReportExportIntegrationTests(TenantTestMixin, TestCase):
    def setUp(self):
        self.setup_tenant_context(name='Export Tenant', slug='export-tenant')

    def test_compile_then_xlsx_contains_real_asset(self):
        from assets.models import Asset, StatusLabel, AssetRole
        from extras.models import ReportTemplate
        from core.reports import compile_report_context

        status = StatusLabel.objects.create(name='Deployed Exp', slug='deployed-exp',
                                             type='deployed', color='28a745')
        role = AssetRole.objects.create(name='Laptop Exp', slug='laptop-exp')
        Asset.objects.create(name='Export Laptop', asset_tag='EXP-XYZ', status=status,
                             asset_role=role, tenant=self.tenant,
                             purchase_cost=1000, currency='EUR')

        tpl = ReportTemplate.objects.create(
            name='Export Rpt', report_type=ReportTemplate.REPORT_TYPE_ASSET_SUMMARY,
            tenant=self.tenant, included_columns=['asset_tag', 'name', 'purchase_cost'],
        )
        headers, rows, *_ = compile_report_context(tpl, active_tenant=self.tenant)

        data = report_xlsx_bytes(headers, rows, sheet_title=tpl.name)
        from openpyxl import load_workbook
        ws = load_workbook(io.BytesIO(data)).active
        all_values = [ws.cell(row=r, column=c).value
                      for r in range(1, ws.max_row + 1) for c in range(1, ws.max_column + 1)]
        self.assertIn('EXP-XYZ', all_values)
        # The EUR purchase cost must NOT be rendered with a '$'.
        money_cells = [v for v in all_values if isinstance(v, str) and ('1,000' in v or '1.000' in v)]
        self.assertTrue(money_cells, 'expected a formatted purchase-cost cell')
        self.assertFalse(any(v.startswith('$') for v in money_cells))
